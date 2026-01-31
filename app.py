from flask import Flask, request, abort
from linebot import LineBotApi, WebhookHandler
from linebot.exceptions import InvalidSignatureError
from linebot.models import MessageEvent, TextMessage, TextSendMessage
from datetime import datetime
import openpyxl
import os

app = Flask(__name__)

LINE_CHANNEL_ACCESS_TOKEN = "qNpwahsG0/l3GyfJ3JEVpvNHkbvgKtcKQRlLgWiWsJMk5IcqRenTu8I94dCg+YwlQNqJJTQa/LXa0MT1jZPgEBATwV2pZ99337xofweLJbrTMzrUSElFTg+lMWc01TAanaTBqtKswWvyqwPP4CTfygdB04t89/1O/w1cDnyilFU="
LINE_CHANNEL_SECRET = "d98969f847034e6ccc0decbc8db405eb"

line_bot_api = LineBotApi(LINE_CHANNEL_ACCESS_TOKEN)
handler = WebhookHandler(LINE_CHANNEL_SECRET)

@app.route("/callback", methods=["POST"])
def callback():
    signature = request.headers["X-Line-Signature"]
    body = request.get_data(as_text=True)

    try:
        handler.handle(body, signature)
    except InvalidSignatureError:
        abort(400)

    return "OK"


@handler.add(MessageEvent, message=TextMessage)
def handle_message(event):
    text = event.message.text.strip()
    parts = text.split()

    item = ""
    amount = 0
    memo = ""

    if len(parts) >= 1:
        item = parts[0]

    if len(parts) >= 2:
        try:
            amount = int(parts[1])
        except:
            amount = 0

    for p in parts:
        if p.startswith("メモ:"):
            memo = p.replace("メモ:", "")

    income = 0
    expense = 0

    if item in ["売上", "収入"]:
        income = amount
    else:
        expense = amount

    # --- 日付を自動で付ける ---
    date_str = datetime.now().strftime("%Y-%m-%d")

    # --- Excel 保存処理 ---
    excel_path = "収支管理.xlsx"

    # ファイルが無ければ作成
    if not os.path.exists(excel_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "dummy"  # 最初のシート
        wb.save(excel_path)

    # ファイルを開く
    wb = openpyxl.load_workbook(excel_path)

    # シート名（例：2026-02）
    sheet_name = date_str[:7]

    # シートが無ければ作成
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(title=sheet_name)
        ws.append(["日付", "費目", "収入", "支出", "メモ"])
    else:
        ws = wb[sheet_name]

    # データを1行追加
    ws.append([date_str, item, income, expense, memo])

    # 保存
    wb.save(excel_path)

    # --- 返信メッセージ ---
    reply_text = (
        f"記録しました！\n"
        f"日付: {date_str}\n"
        f"費目: {item}\n"
        f"収入: {income}\n"
        f"支出: {expense}\n"
        f"メモ: {memo}"
    )

    line_bot_api.reply_message(
        event.reply_token,
        TextSendMessage(text=reply_text)
    )


if __name__ == "__main__":
    app.run(port=5000)
